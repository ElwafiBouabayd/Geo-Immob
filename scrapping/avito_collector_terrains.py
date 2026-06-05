"""
Collecteur d'annonces Avito - Terrains et fermes a vendre a Casablanca
======================================================================
 
Itere sur les 20 combinaisons de filtres (Zoning x Loti x Titre),
pagine chaque combinaison, parse les annonces et produit un Excel.
 
Filtres URL Avito:
    zoning  : 0=Maison Villa, 1=Immeuble, 2=Agricole, 3=Industriel,
              4=Service public
    parted  : true/false  (Loti)
    titled  : true/false  (Titre)
 
Champs collectes:
    - Zoning, Loti, Titre
    - Surface (m2), Quartier, Adresse (= Quartier)
    - Prix, vendeur, date, statut
 
Installation:
    pip install requests beautifulsoup4 lxml openpyxl
 
Usage:
    python avito_collector_terrains.py
 
Sortie:
    avito_casablanca_terrains_fermes.xlsx
    avito_checkpoint_terrains.json (interne, peut etre supprime apres)
"""
 
import re
import time
import json
from pathlib import Path
from urllib.parse import urlencode, unquote
 
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
 
 
# ============================================================
# CONFIGURATION
# ============================================================
 
BASE_URL = "https://www.avito.ma/fr/casablanca/terrains_et_fermes-à_vendre"
AD_URL_SLUG = "/terrains_et_fermes/"
 
ZONINGS = {
    0: "Maison Villa",
    1: "Immeuble",
    2: "Agricole",
    3: "Industriel",
    4: "Service public",
}
PARTED_OPTIONS = {True: "Oui", False: "Non"}     # Loti
TITLED_OPTIONS = {True: "Oui", False: "Non"}     # Titre
 
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}
 
DELAY_BETWEEN_REQUESTS = 1.5  # secondes
OUTPUT_FILE = "avito_casablanca_terrains_fermes.xlsx"
CHECKPOINT_FILE = "avito_checkpoint_terrains.json"
 
 
# ============================================================
# REQUETES HTTP
# ============================================================
 
def build_url(zoning_idx, parted, titled, page=1):
    params = {
        "has_price": "true",
        "parted": "true" if parted else "false",
        "titled": "true" if titled else "false",
        "zoning": zoning_idx,
    }
    if page > 1:
        params["o"] = page
    return BASE_URL + "?" + urlencode(params)
 
 
def fetch_page(url, retries=3):
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=30)
            if resp.status_code == 200:
                return resp.text
            print("  [WARN] HTTP", resp.status_code, "pour", url)
        except requests.RequestException as e:
            print("  [WARN] tentative", attempt + 1, "echec:", e)
        time.sleep(2 ** attempt)
    return None
 
 
# ============================================================
# PARSING
# ============================================================
 
def parse_listing_count(html):
    """Nombre total d'annonces sur la combinaison filtree."""
    soup = BeautifulSoup(html, "lxml")
    h1 = soup.find("h1")
    if h1:
        m = re.search(r":\s*([\d\s]+)\s*annonces", h1.get_text())
        if m:
            return int(m.group(1).replace(" ", ""))
    return 0
 
 
def smart_titlecase(s):
    return " ".join(w[:1].upper() + w[1:] for w in s.split() if w)
 
 
def extract_title_from_url(href):
    if AD_URL_SLUG not in href:
        return ""
    url_path = href.split(AD_URL_SLUG)[-1]
    if url_path.endswith(".htm"):
        url_path = url_path[:-4]
    parts = url_path.rsplit("_", 1)
    if len(parts) == 2 and parts[1].isdigit():
        slug = parts[0]
    else:
        slug = url_path
    return unquote(slug).replace("_", " ").strip()
 
 
def extract_quartier_from_url(href):
    """Quartier = segment entre /fr/ et /terrains_et_fermes/."""
    m = re.search(r"/fr/([^/]+)" + re.escape(AD_URL_SLUG), href)
    if not m:
        return ""
    slug = unquote(m.group(1))
    quartier = slug.replace("_", " ").strip()
    if quartier.lower() == "autre secteur":
        return "Autre secteur"
    return smart_titlecase(quartier)
 
 
def parse_listings(html):
    """Parse toutes les annonces d'une page de resultats."""
    soup = BeautifulSoup(html, "lxml")
    listings = []
    seen_ids = set()
 
    for link in soup.find_all("a", href=True):
        href = link["href"]
 
        # Filtrage: garder uniquement les annonces individuelles
        if AD_URL_SLUG not in href:
            continue
        if "/maroc/" in href:
            continue
        if "terrains_et_fermes-" in href:   # page de listing
            continue
        if "immoneuf" in href:               # promotions immo neuf
            continue
        if not href.endswith(".htm"):
            continue
 
        ad_id_match = re.search(r"_(\d+)\.htm$", href)
        if not ad_id_match:
            continue
        ad_id = ad_id_match.group(1)
        if ad_id in seen_ids:
            continue
        seen_ids.add(ad_id)
 
        text = link.get_text(separator=" ", strip=True)
        text = re.sub(r"\s+", " ", text)
 
        title = extract_title_from_url(href)
        quartier = extract_quartier_from_url(href)
 
        data = {
            "id": ad_id,
            "url": href if href.startswith("http") else "https://www.avito.ma" + href,
            "titre": title,
            "prix_dh": None,
            "mensualite_dh": None,
            "surface_m2": None,
            "quartier": quartier,
            "adresse": quartier,
            "vendeur": "",
            "date_publication": "",
            "statut": "",
        }
 
        # Date publication
        m = re.search(r"il y a \d+\s+\w+", text)
        if m:
            data["date_publication"] = m.group(0)
 
        # Statut
        for tag in ("Premium", "Star", "Vérifié"):
            if tag in text:
                data["statut"] = tag
                break
 
        # Surface
        m = re.search(r"(\d+)\s*m²", text)
        if m:
            data["surface_m2"] = int(m.group(1))
 
        # Mensualite
        m = re.search(r"(\d[\d\s]*)\s*DH\s*/\s*mois", text)
        if m:
            try:
                data["mensualite_dh"] = int(m.group(1).replace(" ", ""))
            except ValueError:
                pass
 
        # Prix
        prix_matches = re.findall(r"(\d[\d\s]*)\s*DH(?!\s*/)", text)
        if prix_matches:
            try:
                data["prix_dh"] = int(prix_matches[0].replace(" ", ""))
            except ValueError:
                pass
 
        # Vendeur (debut, souvent double)
        m = re.match(
            r"^(.+?)(?=\s*il y a|\s*Premium|\s*Star|\s*Vérifié|\s*Terrain|\s*Ferme|\s*\d)",
            text
        )
        if m:
            vendeur_raw = m.group(1).strip()
            half = len(vendeur_raw) // 2
            if half > 0 and vendeur_raw[:half] == vendeur_raw[half:half * 2]:
                vendeur_raw = vendeur_raw[:half].strip()
            if 0 < len(vendeur_raw) < 80:
                data["vendeur"] = vendeur_raw
 
        if data["prix_dh"] is not None and data["prix_dh"] > 0:
            listings.append(data)
 
    return listings
 
 
# ============================================================
# COLLECTE
# ============================================================
 
def load_checkpoint():
    p = Path(CHECKPOINT_FILE)
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return {"done_combos": [], "annonces": {}}
 
 
def save_checkpoint(state):
    Path(CHECKPOINT_FILE).write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
 
 
def collect_combination(zoning_idx, parted, titled):
    zon = ZONINGS[zoning_idx]
    lot = PARTED_OPTIONS[parted]
    tit = TITLED_OPTIONS[titled]
    print("\n  Filtres:", zon, "| Loti:", lot, "| Titre:", tit)
 
    url = build_url(zoning_idx, parted, titled, page=1)
    html = fetch_page(url)
    if not html:
        print("  [ERROR] page 1 inaccessible")
        return []
 
    total = parse_listing_count(html)
    print("    Total annonces:", total)
    if total == 0:
        return []
 
    all_listings = parse_listings(html)
    print("    Page 1:", len(all_listings), "annonces parsees")
 
    nb_pages = max(1, (total + 39) // 40)
 
    for page in range(2, nb_pages + 1):
        time.sleep(DELAY_BETWEEN_REQUESTS)
        url = build_url(zoning_idx, parted, titled, page=page)
        html = fetch_page(url)
        if not html:
            print("    [WARN] page", page, "echec")
            continue
        page_listings = parse_listings(html)
        all_listings.extend(page_listings)
        print("    Page", page, "/", nb_pages, ":", len(page_listings),
              "(cumul:", len(all_listings), ")")
        if len(page_listings) == 0:
            break
 
    for listing in all_listings:
        listing["zoning"] = zon
        listing["loti"] = lot
        listing["titre_bien"] = tit
 
    return all_listings
 
 
def collect_all():
    state = load_checkpoint()
    annonces = state["annonces"]
    done = set(tuple(c) for c in state["done_combos"])
 
    total_combos = len(ZONINGS) * len(PARTED_OPTIONS) * len(TITLED_OPTIONS)
    combo_num = 0
 
    for zoning_idx in ZONINGS:
        for parted in PARTED_OPTIONS:
            for titled in TITLED_OPTIONS:
                combo_num += 1
                key = (zoning_idx, int(parted), int(titled))
                print("\n[" + str(combo_num) + "/" + str(total_combos) + "]",
                      end="")
 
                if key in done:
                    print(" Deja fait, on passe.")
                    continue
 
                listings = collect_combination(zoning_idx, parted, titled)
                for listing in listings:
                    # Ne pas ecraser une annonce deja vue
                    if listing["id"] not in annonces:
                        annonces[listing["id"]] = listing
 
                state["done_combos"].append(list(key))
                state["annonces"] = annonces
                save_checkpoint(state)
                time.sleep(DELAY_BETWEEN_REQUESTS)
 
    print("\n\nCollecte terminee:", len(annonces), "annonces uniques")
    return annonces
 
 
# ============================================================
# EXPORT EXCEL
# ============================================================
 
def export_to_excel(annonces, filename=OUTPUT_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Terrains-Fermes"
 
    headers = [
        "ID", "Titre", "Prix (DH)", "Mensualite (DH/mois)",
        "Surface (m2)", "Quartier", "Adresse",
        "Zoning", "Loti", "Titre",
        "Vendeur", "Statut", "Date publication", "Lien",
    ]
    ws.append(headers)
 
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78",
                              fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
 
    for a in annonces.values():
        ws.append([
            a.get("id", ""),
            a.get("titre", ""),
            a.get("prix_dh"),
            a.get("mensualite_dh"),
            a.get("surface_m2"),
            a.get("quartier", ""),
            a.get("adresse", ""),
            a.get("zoning", ""),
            a.get("loti", ""),
            a.get("titre_bien", ""),
            a.get("vendeur", ""),
            a.get("statut", ""),
            a.get("date_publication", ""),
            a.get("url", ""),
        ])
 
    widths = [12, 45, 14, 16, 12, 22, 22, 16, 8, 8, 22, 10, 18, 60]
    for i, w in enumerate(widths, start=1):
        if i <= 26:
            col_letter = chr(64 + i)
        else:
            col_letter = "A" + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = w
 
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
 
    wb.save(filename)
    print("\nFichier Excel genere:", filename)
    print("  Nombre d'annonces:", len(annonces))
 
 
# ============================================================
# MAIN
# ============================================================
 
def main():
    print("=" * 60)
    print("Collecteur Avito - Terrains et fermes a vendre a Casablanca")
    print("=" * 60)
    nb = len(ZONINGS) * len(PARTED_OPTIONS) * len(TITLED_OPTIONS)
    print("Filtres:", len(ZONINGS), "zoning x",
          len(PARTED_OPTIONS), "loti x", len(TITLED_OPTIONS), "titre =",
          nb, "combinaisons")
    print("Delai entre requetes:", DELAY_BETWEEN_REQUESTS, "s")
    print("Checkpoint:", CHECKPOINT_FILE, "(reprise auto si interrompu)")
    print()
 
    annonces = collect_all()
    export_to_excel(annonces)
    print("\nTermine!")
 
 
if __name__ == "__main__":
    main()
 