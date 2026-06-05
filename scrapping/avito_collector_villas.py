"""
Collecteur d'annonces Avito - Villas/Riad a vendre a Casablanca
================================================================
 
Itere sur les 45 combinaisons de filtres (Condition x Standing x Age),
pagine chaque combinaison, parse les annonces et produit un Excel.
 
Champs collectes:
    - Condition (Neuf, Bon etat, A renover)
    - Standing (Economique, Moyen standing, Haut standing)
    - Age du bien (Moins de 1 an, 1-5 ans, 6-10 ans, 11-20 ans, 21+ ans)
    - Surface (m2)
    - Quartier
    - Adresse (= Quartier, l'adresse complete n'est pas publiee sur la liste)
    - + Prix, chambres, salles de bain, vendeur, date, etc.
 
Installation:
    pip install requests beautifulsoup4 lxml openpyxl
 
Usage:
    python avito_collector_villas.py
 
Le script peut etre interrompu et relance: il reprend automatiquement
grace au fichier checkpoint.
 
Sortie:
    avito_casablanca_villas_riad.xlsx
    avito_checkpoint_villas.json (interne, peut etre supprime apres)
"""
 
import re
import time
import json
from pathlib import Path
from urllib.parse import urlencode, unquote
 
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
 
 
# ============================================================
# CONFIGURATION
# ============================================================
 
# URL de listing (page de resultats) - slug "villas_riad-à_vendre"
BASE_URL = "https://www.avito.ma/fr/casablanca/villas_riad-à_vendre"
 
# Slug present dans les URL des annonces individuelles
# IMPORTANT: la page de resultats utilise "villas_riad" mais les annonces
# individuelles utilisent "villas_et_riads" dans leur URL.
AD_URL_SLUG = "/villas_et_riads/"
 
# Mapping des indices URL d'Avito vers les libelles humains
CONDITIONS = {0: "Neuf", 1: "Bon etat", 2: "A renover"}
STANDINGS = {0: "Economique", 1: "Moyen standing", 2: "Haut standing"}
AGES = {0: "Moins de 1 an", 1: "1-5 ans", 2: "6-10 ans",
        3: "11-20 ans", 4: "21+ ans"}
 
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}
 
DELAY_BETWEEN_REQUESTS = 1.5  # secondes
OUTPUT_FILE = "avito_casablanca_villas_riad.xlsx"
CHECKPOINT_FILE = "avito_checkpoint_villas.json"
 
 
# ============================================================
# REQUETES HTTP
# ============================================================
 
def build_url(condition_idx, standing_idx, age_idx, page=1):
    params = {
        "has_price": "true",
        "property_condition": condition_idx,
        "property_standing": standing_idx,
        "property_age": age_idx,
    }
    if page > 1:
        params["o"] = page
    return BASE_URL + "?" + urlencode(params)
 
 
def fetch_page(url, retries=3):
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=30)
            if resp.status_code == 200:
                return resp.text
            print("  [WARN] HTTP", resp.status_code, "pour", url)
        except requests.RequestException as e:
            print("  [WARN] tentative", attempt + 1, "echec:", e)
        time.sleep(2 ** attempt)
    return None
 
 
# ============================================================
# PARSING
# ============================================================
 
def parse_listing_count(html):
    """Nombre total d'annonces sur la combinaison filtree."""
    soup = BeautifulSoup(html, "lxml")
    h1 = soup.find("h1")
    if h1:
        m = re.search(r":\s*([\d\s]+)\s*annonces", h1.get_text())
        if m:
            return int(m.group(1).replace(" ", ""))
    return 0
 
 
def smart_titlecase(s):
    """Capitalise la 1re lettre de chaque mot sans casser les accents."""
    return " ".join(w[:1].upper() + w[1:] for w in s.split() if w)
 
 
def extract_title_from_url(href):
    """Extrait le titre lisible depuis le slug d'URL."""
    if AD_URL_SLUG not in href:
        return ""
    url_path = href.split(AD_URL_SLUG)[-1]
    if url_path.endswith(".htm"):
        url_path = url_path[:-4]
    parts = url_path.rsplit("_", 1)
    if len(parts) == 2 and parts[1].isdigit():
        slug = parts[0]
    else:
        slug = url_path
    return unquote(slug).replace("_", " ").strip()
 
 
def extract_quartier_from_url(href):
    """Extrait le quartier depuis le chemin /fr/QUARTIER/villas_et_riads/..."""
    m = re.search(r"/fr/([^/]+)" + re.escape(AD_URL_SLUG), href)
    if not m:
        return ""
    slug = unquote(m.group(1))
    quartier = slug.replace("_", " ").strip()
    # Petite normalisation des cas connus
    if quartier.lower() == "autre secteur":
        return "Autre secteur"
    return smart_titlecase(quartier)
 
 
def extract_quartier_from_text(text, title):
    """Backup: extrait quartier depuis le texte agrege.
 
    Pattern: 'dans Casablanca, [QUARTIER][TITLE]' (sans espace entre quartier
    et titre dans la version concatenee).
    """
    if not title:
        return ""
    # Essai 1: pattern direct avec le titre connu
    pat = r"dans Casablanca,\s*(.+?)" + re.escape(title)
    m = re.search(pat, text)
    if m:
        q = m.group(1).strip()
        if 0 < len(q) < 60:
            return q
    # Essai 2: borne sur le 1er mot du titre
    first_word = title.split()[0] if title.split() else ""
    if first_word:
        pat = (r"dans Casablanca,\s+([A-Za-zÀ-ÿ' \-]+?)(?="
               + re.escape(first_word) + ")")
        m = re.search(pat, text)
        if m:
            q = m.group(1).strip()
            if 0 < len(q) < 60:
                return q
    return ""
 
 
def parse_listings(html):
    """Parse toutes les annonces d'une page de resultats."""
    soup = BeautifulSoup(html, "lxml")
    listings = []
    seen_ids = set()
 
    for link in soup.find_all("a", href=True):
        href = link["href"]
 
        # Filtrage: garder uniquement les annonces individuelles villas/riad
        if AD_URL_SLUG not in href:
            continue
        if "/maroc/" in href:
            continue
        if "villas_riad-" in href:        # exclut la page de listing elle-meme
            continue
        if "immoneuf" in href:            # exclut les promotions immo neuf
            continue
        if not href.endswith(".htm"):
            continue
 
        # ID
        ad_id_match = re.search(r"_(\d+)\.htm$", href)
        if not ad_id_match:
            continue
        ad_id = ad_id_match.group(1)
        if ad_id in seen_ids:
            continue
        seen_ids.add(ad_id)
 
        # Texte agrege
        text = link.get_text(separator=" ", strip=True)
        text = re.sub(r"\s+", " ", text)
 
        # Quartier - priorite a l'URL, fallback texte
        title = extract_title_from_url(href)
        quartier = extract_quartier_from_url(href)
        if not quartier:
            quartier = extract_quartier_from_text(text, title)
 
        data = {
            "id": ad_id,
            "url": href if href.startswith("http") else "https://www.avito.ma" + href,
            "titre": title,
            "prix_dh": None,
            "mensualite_dh": None,
            "chambres": None,
            "salles_de_bain": None,
            "surface_m2": None,
            "etage": "",
            "quartier": quartier,
            "adresse": quartier,   # adresse non publiee sur la liste -> = quartier
            "vendeur": "",
            "date_publication": "",
            "statut": "",
        }
 
        # Date publication
        m = re.search(r"il y a \d+\s+\w+", text)
        if m:
            data["date_publication"] = m.group(0)
 
        # Statut
        for tag in ("Premium", "Star", "Vérifié"):
            if tag in text:
                data["statut"] = tag
                break
 
        # Chambres
        m = re.search(r"(\d+)\s*chambres?\b", text)
        if m:
            data["chambres"] = int(m.group(1))
 
        # Salles de bain
        m = re.search(r"(\d+)\s*sdbs?\b", text)
        if m:
            data["salles_de_bain"] = int(m.group(1))
 
        # Surface
        m = re.search(r"(\d+)\s*m²", text)
        if m:
            data["surface_m2"] = int(m.group(1))
 
        # Etage (rare pour villas, mais on garde)
        m = re.search(r"Étage\s+(\d+)", text)
        if m:
            data["etage"] = "Etage " + m.group(1)
        elif "Rez de chaussée" in text:
            data["etage"] = "RDC"
 
        # Mensualite (avant le prix)
        m = re.search(r"(\d[\d\s]*)\s*DH\s*/\s*mois", text)
        if m:
            try:
                data["mensualite_dh"] = int(m.group(1).replace(" ", ""))
            except ValueError:
                pass
 
        # Prix (premier DH non suivi de /mois)
        prix_matches = re.findall(r"(\d[\d\s]*)\s*DH(?!\s*/)", text)
        if prix_matches:
            try:
                data["prix_dh"] = int(prix_matches[0].replace(" ", ""))
            except ValueError:
                pass
 
        # Vendeur (debut du texte, souvent double)
        m = re.match(
            r"^(.+?)(?=\s*il y a|\s*Premium|\s*Star|\s*Vérifié|\s*Villa|\s*Riad|\s*\d)",
            text
        )
        if m:
            vendeur_raw = m.group(1).strip()
            half = len(vendeur_raw) // 2
            if half > 0 and vendeur_raw[:half] == vendeur_raw[half:half * 2]:
                vendeur_raw = vendeur_raw[:half].strip()
            if 0 < len(vendeur_raw) < 80:
                data["vendeur"] = vendeur_raw
 
        if data["prix_dh"] is not None and data["prix_dh"] > 0:
            listings.append(data)
 
    return listings
 
 
# ============================================================
# COLLECTE
# ============================================================
 
def load_checkpoint():
    p = Path(CHECKPOINT_FILE)
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return {"done_combos": [], "annonces": {}}
 
 
def save_checkpoint(state):
    Path(CHECKPOINT_FILE).write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
 
 
def collect_combination(cond_idx, std_idx, age_idx):
    cond = CONDITIONS[cond_idx]
    std = STANDINGS[std_idx]
    age = AGES[age_idx]
    print("\n  Filtres:", cond, "|", std, "|", age)
 
    url = build_url(cond_idx, std_idx, age_idx, page=1)
    html = fetch_page(url)
    if not html:
        print("  [ERROR] page 1 inaccessible")
        return []
 
    total = parse_listing_count(html)
    print("    Total annonces:", total)
    if total == 0:
        return []
 
    all_listings = parse_listings(html)
    print("    Page 1:", len(all_listings), "annonces parsees")
 
    # ~40 annonces utiles par page
    nb_pages = max(1, (total + 39) // 40)
 
    for page in range(2, nb_pages + 1):
        time.sleep(DELAY_BETWEEN_REQUESTS)
        url = build_url(cond_idx, std_idx, age_idx, page=page)
        html = fetch_page(url)
        if not html:
            print("    [WARN] page", page, "echec")
            continue
        page_listings = parse_listings(html)
        all_listings.extend(page_listings)
        print("    Page", page, "/", nb_pages, ":", len(page_listings),
              "(cumul:", len(all_listings), ")")
        if len(page_listings) == 0:
            break
 
    # Tagger les annonces avec leurs attributs
    for listing in all_listings:
        listing["condition"] = cond
        listing["standing"] = std
        listing["age_bien"] = age
 
    return all_listings
 
 
def collect_all():
    state = load_checkpoint()
    annonces = state["annonces"]
    done = set(tuple(c) for c in state["done_combos"])
 
    total_combos = len(CONDITIONS) * len(STANDINGS) * len(AGES)
    combo_num = 0
 
    for cond_idx in CONDITIONS:
        for std_idx in STANDINGS:
            for age_idx in AGES:
                combo_num += 1
                key = (cond_idx, std_idx, age_idx)
                print("\n[" + str(combo_num) + "/" + str(total_combos) + "]", end="")
 
                if key in done:
                    print(" Deja fait, on passe.")
                    continue
 
                listings = collect_combination(cond_idx, std_idx, age_idx)
                for listing in listings:
                    # On ne re-ecrase pas si deja vu avec des filtres differents
                    # (l'annonce garde son 1er triplet condition/standing/age)
                    if listing["id"] not in annonces:
                        annonces[listing["id"]] = listing
 
                state["done_combos"].append(list(key))
                state["annonces"] = annonces
                save_checkpoint(state)
                time.sleep(DELAY_BETWEEN_REQUESTS)
 
    print("\n\nCollecte terminee:", len(annonces), "annonces uniques")
    return annonces
 
 
# ============================================================
# EXPORT EXCEL
# ============================================================
 
def export_to_excel(annonces, filename=OUTPUT_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Villas-Riad"
 
    headers = [
        "ID", "Titre", "Prix (DH)", "Mensualite (DH/mois)",
        "Chambres", "Salles de bain", "Surface (m2)", "Etage",
        "Quartier", "Adresse",
        "Condition", "Standing", "Age du bien",
        "Vendeur", "Statut", "Date publication", "Lien",
    ]
    ws.append(headers)
 
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78",
                              fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
 
    for a in annonces.values():
        ws.append([
            a.get("id", ""),
            a.get("titre", ""),
            a.get("prix_dh"),
            a.get("mensualite_dh"),
            a.get("chambres"),
            a.get("salles_de_bain"),
            a.get("surface_m2"),
            a.get("etage", ""),
            a.get("quartier", ""),
            a.get("adresse", ""),
            a.get("condition", ""),
            a.get("standing", ""),
            a.get("age_bien", ""),
            a.get("vendeur", ""),
            a.get("statut", ""),
            a.get("date_publication", ""),
            a.get("url", ""),
        ])
 
    widths = [12, 45, 14, 16, 10, 10, 12, 14, 22, 22,
              14, 16, 14, 22, 10, 18, 60]
    for i, w in enumerate(widths, start=1):
        # Gestion des colonnes au-dela de Z (>26)
        if i <= 26:
            col_letter = chr(64 + i)
        else:
            col_letter = "A" + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = w
 
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
 
    wb.save(filename)
    print("\nFichier Excel genere:", filename)
    print("  Nombre d'annonces:", len(annonces))
 
 
# ============================================================
# MAIN
# ============================================================
 
def main():
    print("=" * 60)
    print("Collecteur Avito - Villas/Riad a vendre a Casablanca")
    print("=" * 60)
    nb = len(CONDITIONS) * len(STANDINGS) * len(AGES)
    print("Filtres:", len(CONDITIONS), "conditions x",
          len(STANDINGS), "standings x", len(AGES), "ages =", nb,
          "combinaisons")
    print("Delai entre requetes:", DELAY_BETWEEN_REQUESTS, "s")
    print("Checkpoint:", CHECKPOINT_FILE, "(reprise auto si interrompu)")
    print()
 
    annonces = collect_all()
    export_to_excel(annonces)
    print("\nTermine!")
 
 
if __name__ == "__main__":
    main()
 