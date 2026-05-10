"""
Pipeline géomarketing Casablanca — densité POI & proximité d'aménités par zone
==============================================================================

Entrée  : Referentiel_Prix_Casablanca_v2.xlsx  (référentiel DGI, 219 zones)
Sortie  : Referentiel_Prix_Casablanca_ENRICHI.xlsx
          - Feuille « Référentiel Prix Casablanca »  (original, intact)
          - Feuille « Métriques par zone »  (1 ligne par zone, ~30 colonnes)
          - Feuille « Zones géocodées »     (lat/lon, qualité du géocodage)

Étapes
------
1. Lit le référentiel Excel et extrait les zones uniques (CodeZone + Zone déchiffrée)
2. Géocode chaque zone via Nominatim/OSM  (1 requête/seconde, conformément à l'usage policy)
3. Télécharge en bloc tous les POI de la bbox de Casablanca via Overpass
   pour 4 catégories : éducation, banque/ATM, santé, commerces & alimentation
4. Pour chaque zone, calcule :
     - nb_<categorie>_500m  et  nb_<categorie>_1km
     - dist_<categorie>_min_m   (distance au plus proche, en mètres)
     - score_amenites_500m / score_amenites_1km   (somme normalisée)
5. Écrit l'Excel enrichi.

Usage
-----
    pip install requests openpyxl
    python run_geomarketing.py

Les résultats Nominatim/Overpass sont mis en cache dans ./cache/  pour pouvoir
relancer le script sans tout retéléchager.

Auteur : Claude (Cowork)  — Mai 2026
"""

from __future__ import annotations

import csv
import json
import math
import os
import re
import sys
import time
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Configuration
# -----------------------------------------------------------------------------
HERE = Path(__file__).resolve().parent
INPUT_XLSX  = HERE.parent / "Referentiel_Prix_Casablanca_v2.xlsx"
OUTPUT_XLSX = HERE.parent / "Referentiel_Prix_Casablanca_ENRICHI.xlsx"
CACHE_DIR   = HERE / "cache"
CACHE_DIR.mkdir(exist_ok=True)

# Politesse Nominatim : 1 requête/seconde, User-Agent identifiable.
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
OVERPASS_URL  = "https://overpass-api.de/api/interpreter"
USER_AGENT    = "GeoMarketing-Casa/1.0 (contact: elwafibouabayd@gmail.com)"
GEOCODE_DELAY = 1.05  # seconds between Nominatim requests

# Bounding box englobante de la préfecture de Casablanca (large)
# Format Overpass: (south, west, north, east)
CASA_BBOX = (33.45, -7.80, 33.70, -7.45)

# Buffers d'analyse
RADII_M = (500, 1000)   # 500 m et 1 km

# -----------------------------------------------------------------------------
# Points de référence Casablanca (lat, lon)  — utilisés pour les "temps_*"
# -----------------------------------------------------------------------------
# Coordonnées approximatives des principaux pôles d'attraction de Casablanca.
# Ces points servent à estimer un "temps de trajet" depuis chaque zone.
REF_POINTS: dict[str, tuple[float, float]] = {
    "centre":      (33.5731, -7.5898),  # Place Mohammed V / Casa-Voyageurs
    "CFC":         (33.5374, -7.6597),  # Casa Finance City (Anfa)
    "Maarif":      (33.5867, -7.6261),  # Centre Maarif
    "SidiMaarouf": (33.5269, -7.6287),  # Sidi Maarouf
    "port":        (33.6055, -7.6094),  # Port de Casablanca
}

# Hypothèse de vitesse moyenne en circulation urbaine (km/h)
URBAN_SPEED_KMH = 30.0
# Coefficient de détour route / vol d'oiseau
ROAD_FACTOR     = 1.30

# -----------------------------------------------------------------------------
# Catégories d'aménités  ->  filtres Overpass
# -----------------------------------------------------------------------------
# Chaque catégorie liste des paires (key, valeurs) à matcher en OR.
AMENITY_CATEGORIES: dict[str, list[tuple[str, list[str]]]] = {
    "education": [
        ("amenity", ["school", "kindergarten", "college", "university", "language_school"]),
    ],
    "banque": [
        ("amenity", ["bank", "atm"]),
    ],
    "sante": [
        ("amenity", ["hospital", "clinic", "doctors", "pharmacy", "dentist"]),
        ("healthcare", ["hospital", "clinic", "doctor", "pharmacy", "dentist", "centre"]),
    ],
    "commerce": [
        ("shop", ["supermarket", "convenience", "mall", "bakery", "butcher",
                  "greengrocer", "department_store", "general"]),
        ("amenity", ["marketplace", "restaurant", "cafe", "fast_food"]),
    ],
    # ---- Transport (positif) -------------------------------------------------
    "transport": [
        ("highway",          ["bus_stop"]),
        ("railway",          ["station", "halt", "tram_stop"]),
        ("public_transport", ["station", "stop_position", "platform"]),
        ("amenity",          ["taxi", "bus_station", "ferry_terminal"]),
        ("aeroway",          ["aerodrome", "terminal"]),
    ],
    # ---- Environnement / qualité du cadre (positif) --------------------------
    "environnement": [
        ("leisure", ["park", "garden", "playground", "sports_centre",
                     "stadium", "pitch", "swimming_pool", "nature_reserve"]),
        ("natural", ["beach", "wood"]),
        ("landuse", ["forest", "recreation_ground", "grass", "meadow"]),
    ],
    # ---- Nuisances / sources de gêne (négatif) ------------------------------
    "nuisance": [
        ("landuse",  ["industrial", "cemetery", "landfill", "quarry"]),
        ("amenity",  ["fuel", "grave_yard", "waste_disposal", "waste_transfer_station"]),
        ("man_made", ["works", "wastewater_plant", "chimney"]),
        ("power",    ["substation", "plant"]),
    ],
}

# -----------------------------------------------------------------------------
# Catégories de voirie (linéaires)  ->  filtres Overpass (ways uniquement)
# -----------------------------------------------------------------------------
# On calcule la DISTANCE PERPENDICULAIRE de chaque zone à la voie la plus
# proche de chaque catégorie. Les voies sont récupérées avec « out geom » pour
# avoir la géométrie complète (liste de points).
ROAD_CATEGORIES: dict[str, list[str]] = {
    "autoroute":       ["motorway", "motorway_link", "trunk", "trunk_link"],
    "voie_primaire":   ["primary", "primary_link"],
    "voie_secondaire": ["secondary", "secondary_link"],
}

# -----------------------------------------------------------------------------
# Outils
# -----------------------------------------------------------------------------

def haversine_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Distance grand-cercle en mètres."""
    R = 6_371_000.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R * math.asin(math.sqrt(a))


def point_to_segment_m(plat: float, plon: float,
                       alat: float, alon: float,
                       blat: float, blon: float) -> float:
    """Distance d'un point à un segment AB (en mètres) en projection
    équirectangulaire locale — exact à <0.5% à l'échelle de Casablanca."""
    # Projection locale autour de plat
    cos_lat = math.cos(math.radians(plat))
    # Conversion en mètres (origine = point P)
    M_PER_DEG_LAT = 111_320.0
    px, py = 0.0, 0.0
    ax = (alon - plon) * M_PER_DEG_LAT * cos_lat
    ay = (alat - plat) * M_PER_DEG_LAT
    bx = (blon - plon) * M_PER_DEG_LAT * cos_lat
    by = (blat - plat) * M_PER_DEG_LAT
    # Projection scalaire de P sur AB
    abx, aby = bx - ax, by - ay
    apx, apy = px - ax, py - ay
    seg_len2 = abx * abx + aby * aby
    if seg_len2 < 1e-9:                         # A == B
        return math.hypot(px - ax, py - ay)
    t = max(0.0, min(1.0, (apx * abx + apy * aby) / seg_len2))
    cx, cy = ax + t * abx, ay + t * aby
    return math.hypot(px - cx, py - cy)


def point_to_polyline_m(plat: float, plon: float,
                        geom: list[tuple[float, float]]) -> float:
    """Distance min entre un point et une polyligne (liste de (lat, lon))."""
    if len(geom) < 2:
        return haversine_m(plat, plon, geom[0][0], geom[0][1]) if geom else float("inf")
    dmin = float("inf")
    for (alat, alon), (blat, blon) in zip(geom, geom[1:]):
        d = point_to_segment_m(plat, plon, alat, alon, blat, blon)
        if d < dmin:
            dmin = d
    return dmin


def polygon_area_m2(coords: list[tuple[float, float]]) -> float:
    """Aire d'un polygone fermé (liste de (lat, lon)) en m²,
    via projection équirectangulaire locale + formule du lacet (shoelace)."""
    n = len(coords)
    if n < 3:
        return 0.0
    lat0 = sum(c[0] for c in coords) / n
    cos_lat = math.cos(math.radians(lat0))
    M_PER_DEG_LAT = 111_320.0
    pts = [((lon - coords[0][1]) * M_PER_DEG_LAT * cos_lat,
            (lat - coords[0][0]) * M_PER_DEG_LAT)
           for lat, lon in coords]
    s = 0.0
    for i in range(n):
        x1, y1 = pts[i]
        x2, y2 = pts[(i + 1) % n]
        s += x1 * y2 - x2 * y1
    return abs(s) / 2.0


def time_minutes(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Temps de trajet routier estimé (en minutes) entre 2 points :
    distance vol d'oiseau × ROAD_FACTOR ÷ URBAN_SPEED_KMH × 60."""
    d_km = haversine_m(lat1, lon1, lat2, lon2) / 1000.0
    return round(d_km * ROAD_FACTOR / URBAN_SPEED_KMH * 60.0, 1)


def clean_query(label: str) -> str:
    """Normalise la 'Zone déchiffrée' en chaîne géocodable."""
    s = label or ""
    s = re.sub(r"\([^)]*\)", "", s)            # parenthèses
    s = re.split(r"[;,]", s)[0]                # premier élément avant , ou ;
    s = s.replace("Av.", "Avenue").replace("Bd", "Boulevard").replace("Rce", "Résidence")
    s = re.sub(r"\s+", " ", s).strip(" -")
    return s


# -----------------------------------------------------------------------------
# 1. Lecture du référentiel
# -----------------------------------------------------------------------------
def load_zones() -> list[dict]:
    print(f"[1/5] Lecture de {INPUT_XLSX.name}")
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
    ws = wb["Référentiel Prix Casablanca"]
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Préfecture")
    seen, zones = set(), []
    for r in rows[hdr_idx + 1:]:
        if not r[2] or r[2] in seen:
            continue
        seen.add(r[2])
        zones.append({
            "Prefecture":     r[0],
            "Arrondissement": r[1],
            "CodeZone":       r[2],
            "ZoneDechiffree": r[3],
            "QueryShort":     clean_query(r[3] or ""),
        })
    print(f"      {len(zones)} zones uniques.")
    return zones


# -----------------------------------------------------------------------------
# 2. Géocodage Nominatim
# -----------------------------------------------------------------------------
def geocode_zones(zones: list[dict]) -> None:
    cache_file = CACHE_DIR / "geocode.json"
    cache: dict[str, dict] = json.loads(cache_file.read_text("utf-8")) if cache_file.exists() else {}
    print(f"[2/5] Géocodage Nominatim ({len(zones)} zones, cache={len(cache)})")

    sess = requests.Session()
    sess.headers.update({"User-Agent": USER_AGENT, "Accept-Language": "fr"})

    n_new = 0
    for i, z in enumerate(zones, 1):
        code = z["CodeZone"]
        if code in cache:
            z.update(cache[code]); continue

        # Stratégie : on tente 3 formulations, on garde la 1re qui matche dans la bbox de Casa
        candidates = [
            f"{z['QueryShort']}, {z['Arrondissement']}, Casablanca, Maroc",
            f"{z['QueryShort']}, Casablanca, Maroc",
            f"{z['Arrondissement']}, Casablanca, Maroc",
        ]
        best = None
        for q in candidates:
            try:
                r = sess.get(NOMINATIM_URL,
                             params={"q": q, "format": "json", "limit": 1, "countrycodes": "ma"},
                             timeout=15)
                time.sleep(GEOCODE_DELAY)
                if r.status_code != 200:
                    continue
                data = r.json()
                if not data:
                    continue
                lat, lon = float(data[0]["lat"]), float(data[0]["lon"])
                # Garde uniquement si dans la bbox Casa
                if CASA_BBOX[0] <= lat <= CASA_BBOX[2] and CASA_BBOX[1] <= lon <= CASA_BBOX[3]:
                    best = {"lat": lat, "lon": lon,
                            "geocode_query": q,
                            "geocode_quality": "ok" if q == candidates[0] else "fallback",
                            "geocode_display": data[0].get("display_name", "")[:200]}
                    break
            except Exception as e:
                print(f"      ! {code}: {e}"); time.sleep(2)

        if not best:
            best = {"lat": None, "lon": None,
                    "geocode_query": candidates[-1],
                    "geocode_quality": "FAILED",
                    "geocode_display": ""}

        cache[code] = best
        z.update(best)
        n_new += 1

        if i % 10 == 0 or i == len(zones):
            print(f"      {i}/{len(zones)}  (nouveau: {n_new})")
            cache_file.write_text(json.dumps(cache, ensure_ascii=False, indent=2), "utf-8")

    cache_file.write_text(json.dumps(cache, ensure_ascii=False, indent=2), "utf-8")
    ok = sum(1 for z in zones if z.get("lat"))
    print(f"      Géocodés OK : {ok}/{len(zones)}")


# -----------------------------------------------------------------------------
# 3. Téléchargement des POI via Overpass
# -----------------------------------------------------------------------------
def build_overpass_query(category_filters: list[tuple[str, list[str]]], bbox: tuple[float, float, float, float]) -> str:
    s, w, n, e = bbox
    parts = []
    for key, vals in category_filters:
        regex = "|".join(re.escape(v) for v in vals)
        for elem in ("node", "way", "relation"):
            parts.append(f'  {elem}["{key}"~"^({regex})$"]({s},{w},{n},{e});')
    return "[out:json][timeout:120];\n(\n" + "\n".join(parts) + "\n);\nout center tags;"


def fetch_pois() -> dict[str, list[dict]]:
    print(f"[3a/5] Téléchargement POI Overpass (bbox Casa)")
    sess = requests.Session()
    sess.headers.update({"User-Agent": USER_AGENT})
    out: dict[str, list[dict]] = {}
    for cat, filters in AMENITY_CATEGORIES.items():
        cache_file = CACHE_DIR / f"poi_{cat}.json"
        if cache_file.exists():
            out[cat] = json.loads(cache_file.read_text("utf-8"))
            print(f"      {cat:14s} : {len(out[cat])} POI (cache)")
            continue

        q = build_overpass_query(filters, CASA_BBOX)
        for attempt in range(3):
            try:
                r = sess.post(OVERPASS_URL, data={"data": q}, timeout=120)
                r.raise_for_status()
                data = r.json()
                break
            except Exception as e:
                print(f"      ! {cat} tentative {attempt+1}: {e}"); time.sleep(5 + 5*attempt)
        else:
            print(f"      ! {cat}: échec après 3 tentatives, on continue avec liste vide")
            data = {"elements": []}

        pois = []
        for el in data.get("elements", []):
            lat = el.get("lat") or (el.get("center") or {}).get("lat")
            lon = el.get("lon") or (el.get("center") or {}).get("lon")
            if lat is None or lon is None:
                continue
            pois.append({"lat": lat, "lon": lon,
                         "name": (el.get("tags") or {}).get("name", ""),
                         "tags": el.get("tags", {})})
        out[cat] = pois
        cache_file.write_text(json.dumps(pois, ensure_ascii=False), "utf-8")
        print(f"      {cat:14s} : {len(pois)} POI")
        time.sleep(2)  # politesse Overpass
    return out


def fetch_green_areas() -> list[list[tuple[float, float]]]:
    """Télécharge les surfaces vertes (parcs/jardins/forêts) avec leur
    géométrie complète, pour calculer des AIRES (m²) — pas juste des points."""
    cache_file = CACHE_DIR / "green_areas.json"
    if cache_file.exists():
        polys = json.loads(cache_file.read_text("utf-8"))
        print(f"      surfaces vertes  : {len(polys)} polygones (cache)")
        return polys
    print(f"[3c/5] Téléchargement surfaces vertes (avec géométrie)")
    s, w, n, e = CASA_BBOX
    sess = requests.Session()
    sess.headers.update({"User-Agent": USER_AGENT})
    q = (
        f'[out:json][timeout:180];'
        f'('
        f'  way["leisure"~"^(park|garden|nature_reserve|playground)$"]({s},{w},{n},{e});'
        f'  way["landuse"~"^(forest|recreation_ground|grass|meadow|cemetery)$"]({s},{w},{n},{e});'
        f'  way["natural"~"^(wood|grassland|scrub)$"]({s},{w},{n},{e});'
        f'  relation["leisure"~"^(park|garden|nature_reserve)$"]({s},{w},{n},{e});'
        f');'
        f'out geom;'
    )
    data = {"elements": []}
    for attempt in range(3):
        try:
            r = sess.post(OVERPASS_URL, data={"data": q}, timeout=180)
            r.raise_for_status()
            data = r.json()
            break
        except Exception as ex:
            print(f"      ! green_areas tentative {attempt+1}: {ex}")
            time.sleep(5 + 5 * attempt)
    polys: list[list[tuple[float, float]]] = []
    for el in data.get("elements", []):
        geom = el.get("geometry") or []
        if len(geom) >= 3:
            polys.append([(g["lat"], g["lon"]) for g in geom])
    cache_file.write_text(json.dumps(polys, ensure_ascii=False), "utf-8")
    print(f"      surfaces vertes  : {len(polys)} polygones")
    time.sleep(2)
    return polys


def fetch_coastline() -> list[list[tuple[float, float]]]:
    """Télécharge la ligne de côte Atlantique (natural=coastline) pour le
    calcul de dist_mer."""
    cache_file = CACHE_DIR / "coastline.json"
    if cache_file.exists():
        polys = json.loads(cache_file.read_text("utf-8"))
        print(f"      ligne de côte    : {len(polys)} segments (cache)")
        return polys
    print(f"[3d/5] Téléchargement ligne de côte")
    s, w, n, e = CASA_BBOX
    sess = requests.Session()
    sess.headers.update({"User-Agent": USER_AGENT})
    q = (f'[out:json][timeout:120];'
         f'way["natural"="coastline"]({s},{w},{n},{e});'
         f'out geom;')
    data = {"elements": []}
    for attempt in range(3):
        try:
            r = sess.post(OVERPASS_URL, data={"data": q}, timeout=120)
            r.raise_for_status()
            data = r.json()
            break
        except Exception as ex:
            print(f"      ! coastline tentative {attempt+1}: {ex}")
            time.sleep(5 + 5 * attempt)
    polys: list[list[tuple[float, float]]] = []
    for el in data.get("elements", []):
        geom = el.get("geometry") or []
        if len(geom) >= 2:
            polys.append([(g["lat"], g["lon"]) for g in geom])
    cache_file.write_text(json.dumps(polys, ensure_ascii=False), "utf-8")
    print(f"      ligne de côte    : {len(polys)} segments")
    time.sleep(2)
    return polys


def fetch_roads() -> dict[str, list[list[tuple[float, float]]]]:
    """Télécharge les voies (ways) avec géométrie pour chaque catégorie de
    voirie. Retourne {categorie: [polyligne_1, polyligne_2, ...]} où chaque
    polyligne est une liste de (lat, lon)."""
    print(f"[3b/5] Téléchargement voirie Overpass (autoroutes / primaires / secondaires)")
    s, w, n, e = CASA_BBOX
    sess = requests.Session()
    sess.headers.update({"User-Agent": USER_AGENT})
    out: dict[str, list[list[tuple[float, float]]]] = {}
    for cat, hw_values in ROAD_CATEGORIES.items():
        cache_file = CACHE_DIR / f"roads_{cat}.json"
        if cache_file.exists():
            out[cat] = json.loads(cache_file.read_text("utf-8"))
            print(f"      {cat:18s} : {len(out[cat])} voies (cache)")
            continue
        regex = "|".join(re.escape(v) for v in hw_values)
        q = (f'[out:json][timeout:120];'
             f'way["highway"~"^({regex})$"]({s},{w},{n},{e});'
             f'out geom;')
        for attempt in range(3):
            try:
                r = sess.post(OVERPASS_URL, data={"data": q}, timeout=120)
                r.raise_for_status()
                data = r.json()
                break
            except Exception as ex:
                print(f"      ! {cat} tentative {attempt+1}: {ex}"); time.sleep(5 + 5*attempt)
        else:
            data = {"elements": []}
        polylines = []
        for el in data.get("elements", []):
            geom = el.get("geometry") or []
            if len(geom) >= 2:
                polylines.append([(g["lat"], g["lon"]) for g in geom])
        out[cat] = polylines
        cache_file.write_text(json.dumps(polylines, ensure_ascii=False), "utf-8")
        print(f"      {cat:18s} : {len(polylines)} voies")
        time.sleep(2)
    return out


# -----------------------------------------------------------------------------
# 4. Métriques par zone
# -----------------------------------------------------------------------------
# Liste des variables géomarketing dérivées (noms cibles métier).
# On les centralise ici pour garantir que chaque zone a TOUTES les colonnes,
# même quand le géocodage échoue (valeurs None).
GEOMARKETING_VARS: list[str] = [
    # voirie
    "dist_boulevard_principal", "nb_axes_500m", "centralite_routiere",
    # transport
    "dist_tram", "nb_stations_1km", "temps_transport_centre",
    # temps vers pôles
    "temps_CFC", "temps_Maarif", "temps_SidiMaarouf", "temps_port",
    # éducation
    "nb_ecoles_1km", "dist_ecole", "densite_education",
    # santé
    "nb_sante_1km", "dist_clinique",
    # commerce
    "nb_restaurants_500m", "nb_commerces_1km",
    # banque
    "nb_banques_1km", "nb_GAB_1km",
    # environnement
    "dist_mer", "dist_parc", "surface_verte_1km",
    # nuisance
    "nb_nuisance",
]


def _filter_by_tags(plist: list[dict],
                    *match: tuple[str, set]) -> list[dict]:
    """Garde les POI dont au moins un (tag_key, tag_value) appartient à match."""
    out = []
    for p in plist:
        tags = p.get("tags") or {}
        for k, vs in match:
            if tags.get(k) in vs:
                out.append(p)
                break
    return out


def compute_metrics(zones: list[dict],
                    pois: dict[str, list[dict]],
                    roads: dict[str, list[list[tuple[float, float]]]],
                    green_areas: list[list[tuple[float, float]]],
                    coastline: list[list[tuple[float, float]]]) -> list[dict]:
    print(f"[4/5] Calcul des métriques 500m / 1km + accessibilité voirie")
    res = []
    for z in zones:
        row = {**z}
        if z.get("lat") is None:
            for cat in AMENITY_CATEGORIES:
                for r_m in RADII_M:
                    row[f"nb_{cat}_{r_m}m"] = None
                row[f"dist_{cat}_min_m"] = None
            for rcat in ROAD_CATEGORIES:
                row[f"dist_{rcat}_min_m"] = None
            row["score_amenites_500m"] = None
            row["score_amenites_1km"]  = None
            row["score_nuisance_500m"] = None
            row["score_accessibilite"] = None
            for v in GEOMARKETING_VARS:
                row[v] = None
            res.append(row); continue

        lat0, lon0 = z["lat"], z["lon"]
        cos_lat = math.cos(math.radians(lat0))
        # bbox rapide pour préfiltrer les POIs (rayon max + 200m de marge)
        max_r = max(RADII_M) + 200
        dlat = max_r / 111_320
        dlon = max_r / (111_320 * cos_lat)
        bb = (lat0 - dlat, lat0 + dlat, lon0 - dlon, lon0 + dlon)

        # ---- POIs ponctuels --------------------------------------------------
        for cat, plist in pois.items():
            counts = {r_m: 0 for r_m in RADII_M}
            dmin = float("inf")
            for p in plist:
                if not (bb[0] <= p["lat"] <= bb[1] and bb[2] <= p["lon"] <= bb[3]):
                    continue
                d = haversine_m(lat0, lon0, p["lat"], p["lon"])
                if d < dmin:
                    dmin = d
                for r_m in RADII_M:
                    if d <= r_m:
                        counts[r_m] += 1
            for r_m in RADII_M:
                row[f"nb_{cat}_{r_m}m"] = counts[r_m]
            row[f"dist_{cat}_min_m"] = round(dmin) if dmin < float("inf") else None

        # ---- Voirie (linéaire) ----------------------------------------------
        # bbox plus large pour les voies (on peut être à plusieurs km)
        big_dlat = 5000 / 111_320
        big_dlon = 5000 / (111_320 * cos_lat)
        big_bb = (lat0 - big_dlat, lat0 + big_dlat, lon0 - big_dlon, lon0 + big_dlon)
        for rcat, polylines in roads.items():
            dmin = float("inf")
            for poly in polylines:
                # Préfiltre : si AUCUN point de la polyligne n'est dans la bbox
                # élargie, on saute (gros gain de perf).
                if not any(big_bb[0] <= la <= big_bb[1] and big_bb[2] <= lo <= big_bb[3]
                           for la, lo in poly):
                    continue
                d = point_to_polyline_m(lat0, lon0, poly)
                if d < dmin:
                    dmin = d
                    if dmin < 5:        # déjà sur la voie, inutile de continuer
                        break
            row[f"dist_{rcat}_min_m"] = round(dmin) if dmin < float("inf") else None

        # ---- Scores composites ---------------------------------------------
        # Aménités positives : poids selon importance perçue pour le résidentiel
        weights_pos = {"education": 1.5, "sante": 1.5, "banque": 1.0,
                       "commerce": 1.0, "transport": 1.2, "environnement": 1.0}
        for r_m in RADII_M:
            s = sum(w * (row.get(f"nb_{c}_{r_m}m") or 0)
                    for c, w in weights_pos.items())
            row[f"score_amenites_{'500m' if r_m == 500 else '1km'}"] = round(s, 1)

        # Nuisance : nb de POI nuisibles dans 500m + bonus si autoroute < 200m
        nuis = (row.get("nb_nuisance_500m") or 0) * 1.0
        d_auto = row.get("dist_autoroute_min_m")
        if d_auto is not None and d_auto < 200:
            nuis += (200 - d_auto) / 100      # +0 à +2 selon proximité
        row["score_nuisance_500m"] = round(nuis, 1)

        # Accessibilité : décroissance avec distance à voie primaire/secondaire
        # max 10 si on est sur la voie, 0 si > 1500m
        def road_score(d):
            if d is None:        return 0.0
            if d <= 50:          return 10.0
            if d >= 1500:        return 0.0
            return round(10 * (1500 - d) / 1450, 2)
        s_acc = (1.0 * road_score(row.get("dist_voie_primaire_min_m")) +
                 0.6 * road_score(row.get("dist_voie_secondaire_min_m")) +
                 0.4 * road_score(row.get("dist_autoroute_min_m")))
        row["score_accessibilite"] = round(s_acc, 1)

        # ============================================================
        # Variables géomarketing dérivées (noms cibles métier)
        # ============================================================

        # ---- 1) Voirie -----------------------------------------------------
        # Boulevard principal ≈ voie primaire (Bd Zerktouni, Bd d'Anfa, etc.)
        row["dist_boulevard_principal"] = row.get("dist_voie_primaire_min_m")

        # nb_axes_500m : nb de voies majeures (autoroute+primaire+secondaire)
        # passant à <500 m de la zone
        nb_axes = 0
        # bbox élargie de 600 m pour préfiltrer
        ax_dlat = 600 / 111_320
        ax_dlon = 600 / (111_320 * cos_lat)
        ax_bb = (lat0 - ax_dlat, lat0 + ax_dlat,
                 lon0 - ax_dlon, lon0 + ax_dlon)
        for _rcat, polylines in roads.items():
            for poly in polylines:
                if not any(ax_bb[0] <= la <= ax_bb[1] and ax_bb[2] <= lo <= ax_bb[3]
                           for la, lo in poly):
                    continue
                if point_to_polyline_m(lat0, lon0, poly) <= 500:
                    nb_axes += 1
        row["nb_axes_500m"] = nb_axes

        # centralite_routiere : score [0..10+] combinant proximité aux voies
        # et nombre d'axes à proximité
        def _decay(d, max_d):
            if d is None or d >= max_d:
                return 0.0
            return (max_d - d) / max_d
        cr = (3.0 * _decay(row.get("dist_voie_primaire_min_m"),   1500) +
              2.0 * _decay(row.get("dist_voie_secondaire_min_m"), 1000) +
              1.5 * _decay(row.get("dist_autoroute_min_m"),       2000) +
              0.10 * min(nb_axes, 30))
        row["centralite_routiere"] = round(cr, 2)

        # ---- 2) Transport --------------------------------------------------
        transport_pois = pois.get("transport", [])
        trams = _filter_by_tags(transport_pois,
                                ("railway", {"tram_stop"}),
                                ("public_transport", {"tram_stop"}))
        stations = _filter_by_tags(
            transport_pois,
            ("railway",          {"station", "halt", "tram_stop"}),
            ("public_transport", {"station"}),
            ("amenity",          {"bus_station", "ferry_terminal"}),
        )

        dmin_tram = float("inf")
        for p in trams:
            d = haversine_m(lat0, lon0, p["lat"], p["lon"])
            if d < dmin_tram:
                dmin_tram = d
        row["dist_tram"] = round(dmin_tram) if dmin_tram < float("inf") else None

        nb_st = sum(1 for p in stations
                    if haversine_m(lat0, lon0, p["lat"], p["lon"]) <= 1000)
        row["nb_stations_1km"] = nb_st

        # ---- 3) Temps de trajet vers pôles d'attraction --------------------
        # Note : estimation simple (vol d'oiseau × 1.3 / 30 km/h × 60).
        # Pas de routage réel — pour cela, brancher OSRM/Mapbox plus tard.
        for name, (rlat, rlon) in REF_POINTS.items():
            t = time_minutes(lat0, lon0, rlat, rlon)
            if name == "centre":
                row["temps_transport_centre"] = t
            else:
                row[f"temps_{name}"] = t

        # ---- 4) Éducation --------------------------------------------------
        edu_pois = pois.get("education", [])
        ecoles = _filter_by_tags(
            edu_pois,
            ("amenity", {"school", "kindergarten", "college", "university"}),
        )
        dmin_ecole = float("inf")
        nb_ecoles_1km = 0
        for p in ecoles:
            d = haversine_m(lat0, lon0, p["lat"], p["lon"])
            if d < dmin_ecole:
                dmin_ecole = d
            if d <= 1000:
                nb_ecoles_1km += 1
        row["nb_ecoles_1km"] = nb_ecoles_1km
        row["dist_ecole"]    = round(dmin_ecole) if dmin_ecole < float("inf") else None
        # densité = nb_ecoles / aire du disque 1km (= π km²)
        row["densite_education"] = round(nb_ecoles_1km / math.pi, 2)

        # ---- 5) Santé ------------------------------------------------------
        # Alias direct (le 1km existe déjà sous nb_sante_1000m)
        row["nb_sante_1km"] = row.get("nb_sante_1000m")
        sante_pois = pois.get("sante", [])
        cliniques = _filter_by_tags(
            sante_pois,
            ("amenity",    {"clinic", "hospital"}),
            ("healthcare", {"clinic", "hospital", "centre"}),
        )
        dmin_cl = float("inf")
        for p in cliniques:
            d = haversine_m(lat0, lon0, p["lat"], p["lon"])
            if d < dmin_cl:
                dmin_cl = d
        row["dist_clinique"] = round(dmin_cl) if dmin_cl < float("inf") else None

        # ---- 6) Commerce / Restauration ------------------------------------
        com_pois = pois.get("commerce", [])
        restos = _filter_by_tags(
            com_pois,
            ("amenity", {"restaurant", "cafe", "fast_food"}),
        )
        nb_resto_500 = sum(1 for p in restos
                           if haversine_m(lat0, lon0, p["lat"], p["lon"]) <= 500)
        row["nb_restaurants_500m"] = nb_resto_500

        # nb_commerces_1km : tous les "shops" (excluant les restaurants/cafés)
        restos_ids = {id(p) for p in restos}
        nb_com_1km = 0
        for p in com_pois:
            if id(p) in restos_ids:
                continue
            if haversine_m(lat0, lon0, p["lat"], p["lon"]) <= 1000:
                nb_com_1km += 1
        row["nb_commerces_1km"] = nb_com_1km

        # ---- 7) Banque / GAB ----------------------------------------------
        banq_pois = pois.get("banque", [])
        banques = _filter_by_tags(banq_pois, ("amenity", {"bank"}))
        gabs    = _filter_by_tags(banq_pois, ("amenity", {"atm"}))
        row["nb_banques_1km"] = sum(
            1 for p in banques
            if haversine_m(lat0, lon0, p["lat"], p["lon"]) <= 1000
        )
        row["nb_GAB_1km"] = sum(
            1 for p in gabs
            if haversine_m(lat0, lon0, p["lat"], p["lon"]) <= 1000
        )

        # ---- 8) Environnement / cadre de vie -------------------------------
        # dist_mer : distance à la ligne de côte (natural=coastline)
        dmin_mer = float("inf")
        # bbox élargie de 30 km pour la côte (Casa fait ~30 km de côte)
        for poly in coastline:
            # Préfiltre : si AUCUN sommet du segment côtier n'est à <30 km,
            # on saute (très grossier mais la côte est continue).
            if not any(haversine_m(lat0, lon0, la, lo) < 30000 for la, lo in poly):
                continue
            d = point_to_polyline_m(lat0, lon0, poly)
            if d < dmin_mer:
                dmin_mer = d
        row["dist_mer"] = round(dmin_mer) if dmin_mer < float("inf") else None

        # dist_parc : distance au plus proche parc/jardin
        env_pois = pois.get("environnement", [])
        parcs = _filter_by_tags(
            env_pois,
            ("leisure", {"park", "garden", "playground", "nature_reserve"}),
        )
        dmin_parc = float("inf")
        for p in parcs:
            d = haversine_m(lat0, lon0, p["lat"], p["lon"])
            if d < dmin_parc:
                dmin_parc = d
        row["dist_parc"] = round(dmin_parc) if dmin_parc < float("inf") else None

        # surface_verte_1km : aire totale (m²) des polygones verts dont le
        # centroïde est à moins de 1 km de la zone.
        surf = 0.0
        for poly_coords in green_areas:
            # Filtre rapide : si tous les sommets sont à >2 km, skip.
            if all(haversine_m(lat0, lon0, la, lo) > 2000
                   for la, lo in poly_coords):
                continue
            n_pts = len(poly_coords)
            clat = sum(c[0] for c in poly_coords) / n_pts
            clon = sum(c[1] for c in poly_coords) / n_pts
            if haversine_m(lat0, lon0, clat, clon) <= 1000:
                surf += polygon_area_m2(poly_coords)
        row["surface_verte_1km"] = round(surf)

        # ---- 9) Nuisance ---------------------------------------------------
        # Total des sources de nuisance dans 1 km
        row["nb_nuisance"] = row.get("nb_nuisance_1000m") or 0

        res.append(row)
    return res


# -----------------------------------------------------------------------------
# 5. Écriture Excel enrichi
# -----------------------------------------------------------------------------
def write_output(zones_with_metrics: list[dict]) -> None:
    print(f"[5/5] Écriture {OUTPUT_XLSX.name}")
    # On part du fichier source pour conserver toutes les feuilles intactes
    wb = openpyxl.load_workbook(INPUT_XLSX)

    # ---- Feuille Métriques par zone ----
    if "Métriques par zone" in wb.sheetnames:
        del wb["Métriques par zone"]
    ws = wb.create_sheet("Métriques par zone", 1)

    cats = list(AMENITY_CATEGORIES.keys())
    rcats = list(ROAD_CATEGORIES.keys())
    headers = ["CodeZone", "Préfecture", "Arrondissement", "Zone déchiffrée",
               "lat", "lon", "geocode_quality"]
    for cat in cats:
        for r_m in RADII_M:
            headers.append(f"nb_{cat}_{r_m}m")
        headers.append(f"dist_{cat}_min_m")
    for rcat in rcats:
        headers.append(f"dist_{rcat}_min_m")
    headers += ["score_amenites_500m", "score_amenites_1km",
                "score_nuisance_500m", "score_accessibilite"]
    # Variables géomarketing métier (dérivées)
    headers += list(GEOMARKETING_VARS)

    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "E2"

    for z in zones_with_metrics:
        ws.append([z.get(h.replace("Préfecture", "Prefecture")
                          .replace("Zone déchiffrée", "ZoneDechiffree"))
                   if h in ("Préfecture", "Zone déchiffrée") else z.get(h)
                   for h in headers])

    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(11, min(28, len(h) + 2))

    # ---- Feuille Zones géocodées (debug) ----
    if "Zones géocodées" in wb.sheetnames:
        del wb["Zones géocodées"]
    ws2 = wb.create_sheet("Zones géocodées")
    ws2.append(["CodeZone", "Zone déchiffrée", "Query utilisé", "lat", "lon",
                "Qualité", "Display name (Nominatim)"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = header_fill
    for z in zones_with_metrics:
        ws2.append([z["CodeZone"], z["ZoneDechiffree"], z.get("geocode_query"),
                    z.get("lat"), z.get("lon"),
                    z.get("geocode_quality"), z.get("geocode_display")])
    ws2.freeze_panes = "A2"
    for i, w in enumerate([12, 50, 50, 11, 11, 12, 60], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUTPUT_XLSX)
    print(f"      OK -> {OUTPUT_XLSX}")


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main():
    if not INPUT_XLSX.exists():
        sys.exit(f"Fichier introuvable: {INPUT_XLSX}")
    zones = load_zones()
    geocode_zones(zones)
    pois        = fetch_pois()
    roads       = fetch_roads()
    green_areas = fetch_green_areas()
    coastline   = fetch_coastline()
    enriched = compute_metrics(zones, pois, roads, green_areas, coastline)
    write_output(enriched)
    ok = sum(1 for z in enriched if z.get("lat"))
    fail = len(enriched) - ok
    total_pois  = sum(len(v) for v in pois.values())
    total_roads = sum(len(v) for v in roads.values())
    print()
    print("=" * 60)
    print(f"Termine. {ok} zones geocodees, {fail} echecs.")
    print(f"POI Casablanca telecharges : {total_pois}")
    print(f"Voies telechargees         : {total_roads}")
    print(f"Surfaces vertes            : {len(green_areas)} polygones")
    print(f"Segments de cote           : {len(coastline)}")
    print(f"Sortie : {OUTPUT_XLSX}")
    print("=" * 60)


if __name__ == "__main__":
    main()
