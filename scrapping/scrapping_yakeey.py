"""
Yakeey Casablanca - Scraper complet des annonces immobilières
==============================================================

Récupère TOUTES les annonces de vente à Casablanca depuis https://yakeey.com,
y compris les détails (prix, surface, pièces, équipements, GPS) puis génère
un fichier Excel structuré.

Stratégie :
- Le site est en Next.js. Le HTML contient un <script id="__NEXT_DATA__">
  qui embarque toutes les données structurées (listings + détails). C'est
  beaucoup plus fiable que de parser le DOM.

Prérequis :
    pip install requests beautifulsoup4 openpyxl tqdm

Utilisation :
    python yakeey_scraper.py
"""

import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from tqdm import tqdm

# ============================ Configuration ============================

BASE_URL = "https://yakeey.com/fr-ma/achat/biens/casablanca"
OUTPUT_DIR = Path(__file__).parent / "yakeey_data"
OUTPUT_DIR.mkdir(exist_ok=True)
LISTINGS_JSON = OUTPUT_DIR / "listings.json"
DETAILS_JSON = OUTPUT_DIR / "details.json"
EXCEL_FILE = OUTPUT_DIR / "annonces_yakeey_casablanca.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

MAX_WORKERS = 6      # parallélisme des requêtes (respect du serveur)
REQUEST_DELAY = 0.3  # délai entre requêtes pour éviter rate-limit
RETRY = 3            # nombre de tentatives par requête

# ---------------------------------------------------------------------
# STRATEGIE MULTI-VUES (depuis nov. 2025)
# ---------------------------------------------------------------------
# Probleme : la pagination Yakeey "?page=N" est buggee. totalElements
# annonce 1251 mais seulement ~529 IDs uniques sont reellement exposes :
# chaque bien apparait sur jusqu'a 4 pages consecutives (fenetre
# glissante, probablement un tri non-stable cote backend).
#
# Solution : on attaque la liste sous PLUSIEURS angles. Chaque "vue"
# (combinaison sort + filtres) a un pattern d'overlap different, donc
# l'UNION de toutes les vues maximise les IDs uniques.
#
# Vues testees :
#   - Tris : par defaut, newest, oldest, priceAsc, priceDesc
#   - Filtres listingType : PROPERTY seul, PROGRAM seul
#   - Filtres propertyCategory : FLAT, VILLA, OFFICE, LAND, ...
#   - Tranches de prix : on coupe [0..30M] en buckets pour reduire la
#     taille de chaque dataset (et donc l'overlap).
# ---------------------------------------------------------------------

# Liste des parametres de tri testes. On essaie plusieurs variantes
# (camelCase et kebab-case) car on ne sait pas exactement ce que
# Yakeey accepte. Les valeurs non reconnues sont juste ignorees.
SORT_VARIANTS = [
    None,  # tri par defaut
    "newest", "oldest",
    "priceAsc", "priceDesc",
    "price-asc", "price-desc",
    "dateAsc", "dateDesc",
    "date-asc", "date-desc",
    "surfaceAsc", "surfaceDesc",
]

LISTING_TYPES = [None, "PROPERTY", "PROGRAM"]

# Codes connus de propertyCategory chez Yakeey (vu dans le payload RSC)
PROPERTY_CATEGORIES = [
    None, "FLAT", "VILLA", "HOUSE", "OFFICE", "LAND", "RIAD",
    "COMMERCIAL", "COMMERCIAL_BUILDING", "BUILDING", "FARM", "WAREHOUSE",
]

# Tranches de prix (DH). On part de 0 jusqu'a une borne tres haute,
# avec des buckets serres dans la zone ou il y a le plus d'annonces.
PRICE_BUCKETS = [
    (0,       500_000),
    (500_000, 1_000_000),
    (1_000_000, 1_500_000),
    (1_500_000, 2_000_000),
    (2_000_000, 2_500_000),
    (2_500_000, 3_000_000),
    (3_000_000, 4_000_000),
    (4_000_000, 5_000_000),
    (5_000_000, 7_000_000),
    (7_000_000, 10_000_000),
    (10_000_000, 15_000_000),
    (15_000_000, 25_000_000),
    (25_000_000, 100_000_000),
]


# ============================ Helpers HTTP ============================
#
# IMPORTANT : Yakeey/Cloudflare stocke l'etat de pagination dans un cookie
# de session. Si plusieurs threads partagent la meme requests.Session(),
# les requetes paralleles se brouillent et renvoient toutes la meme page.
# Solution : une nouvelle session sans cookies persistants par requete.


def fetch(url: str) -> str | None:
    """GET avec retry et back-off exponentiel, sans cookies persistants."""
    for attempt in range(RETRY):
        try:
            # Session neuve a chaque tentative pour eviter les collisions
            # de cookies entre threads paralleles.
            r = requests.get(url, headers=HEADERS, timeout=30)
            if r.status_code == 200:
                return r.text
            if r.status_code in (429, 503):
                time.sleep(2 ** attempt)
                continue
            print(f"[!] {url} -> HTTP {r.status_code}")
            return None
        except requests.RequestException as e:
            if attempt == RETRY - 1:
                print(f"[!] {url} -> {e}")
                return None
            time.sleep(2 ** attempt)
    return None


def extract_next_data(html: str) -> dict | None:
    """Ancien format (Pages Router). Conservé en fallback."""
    soup = BeautifulSoup(html, "lxml")
    tag = soup.find("script", id="__NEXT_DATA__")
    if tag and tag.string:
        try:
            return json.loads(tag.string)
        except json.JSONDecodeError:
            return None
    return None


def _extract_balanced_json(text: str, start: int) -> str | None:
    """Extrait l'objet JSON commençant à `start` (doit pointer sur '{')."""
    if start >= len(text) or text[start] != "{":
        return None
    depth = 0
    in_str = False
    esc = False
    for j in range(start, len(text)):
        ch = text[j]
        if esc:
            esc = False
            continue
        if ch == "\\":
            esc = True
            continue
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return text[start:j + 1]
    return None


_RSC_PUSH_RE = re.compile(
    r'self\.__next_f\.push\(\[1,"((?:\\.|[^"\\])*)"\]\)', re.DOTALL
)


def extract_rsc_initial_data(html: str) -> dict | None:
    """
    Nouveau format Next.js App Router (RSC streaming).
    Les données sont concaténées dans plusieurs self.__next_f.push([1, "..."]).
    On cherche la clé "initialData":{...} et on en extrait l'objet JSON.
    """
    chunks = _RSC_PUSH_RE.findall(html)
    if not chunks:
        return None
    # Décode les échappements JS de chaque chunk puis concatène
    full = ""
    for c in chunks:
        try:
            full += c.encode("utf-8").decode("unicode_escape")
        except Exception:
            full += c
    key = '"initialData":'
    idx = full.find(key)
    if idx < 0:
        return None
    start = idx + len(key)
    raw = _extract_balanced_json(full, start)
    if not raw:
        return None
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return None


# ============================ Listing pages ============================

def parse_listing_card(item: dict) -> dict | None:
    """
    Convertit un objet 'listing' du nouveau schéma Yakeey (App Router RSC).

    Deux types gérés :
      - PROPERTY : annonce d'un bien unique
          price, area, rooms, bathrooms, internalAddresses, ...
      - PROGRAM : programme immobilier neuf (plusieurs lots)
          priceMin/priceMax, areaMin/areaMax, roomsMin/roomsMax, cities, name, ...
    """
    if not item:
        return None

    listing_type = item.get("listingType") or ""
    is_program = listing_type == "PROGRAM"

    # Référence publique
    ref = (
        item.get("userRef")
        or item.get("reference")
        or item.get("ref")
        or item.get("id")
        or ""
    )

    # Type de bien
    if is_program:
        # Categorie principale : FLAT, VILLA... (depuis propertyCategories ou propertyCategoryCounts)
        cat = ""
        cats = item.get("propertyCategories") or []
        if cats and isinstance(cats[0], dict):
            cat = cats[0].get("code") or ""
        if not cat:
            counts = item.get("propertyCategoryCounts") or {}
            if counts:
                cat = max(counts, key=counts.get)
        type_bien = f"PROGRAMME ({cat})" if cat else "PROGRAMME"
    else:
        type_bien = (
            item.get("propertyType")
            or item.get("propertyCategory")
            or item.get("typeBien")
            or item.get("type")
            or ""
        )

    # Adresse : structure différente selon type
    quartier = ""
    district = ""
    ville = "Casablanca"
    if is_program:
        cities = item.get("cities") or []
        if cities and isinstance(cities[0], dict):
            ville = cities[0].get("name") or ville
            nbhs = cities[0].get("neighborhoods") or []
            primary_nbh = next((n for n in nbhs if n.get("isPrimary")), nbhs[0] if nbhs else {})
            quartier = primary_nbh.get("name") or ""
    else:
        addrs = item.get("internalAddresses") or []
        primary = next((a for a in addrs if a.get("isPrimary")), addrs[0] if addrs else {})
        quartier = (
            primary.get("neighborhood")
            or primary.get("district")
            or item.get("neighborhood")
            or item.get("quartier")
            or ""
        )
        district = primary.get("district") or ""
        ville = (
            primary.get("city")
            or item.get("city")
            or "Casablanca"
        )

    # Prix : pour les programmes, on prend le min comme valeur principale
    if is_program:
        price = item.get("priceMin") or item.get("priceMax") or 0
    else:
        price = (
            item.get("price")
            or item.get("prix")
            or item.get("sellingPrice")
            or 0
        )

    # GPS : "location" est [lng, lat] dans le nouveau schéma
    lat = None
    lng = None
    loc = item.get("location")
    if isinstance(loc, list) and len(loc) >= 2:
        lng, lat = loc[0], loc[1]
    elif isinstance(loc, dict):
        lat = loc.get("latitude") or loc.get("lat")
        lng = loc.get("longitude") or loc.get("lng")
    if lat is None:
        lat = item.get("latitude")
    if lng is None:
        lng = item.get("longitude")

    # URL de l'annonce : pas de slug dans la carte. On utilise l'id ou userRef.
    slug = item.get("slug") or ""
    if slug:
        url = item.get("url") or f"https://yakeey.com/fr-ma/{slug}"
    elif item.get("userRef"):
        url = f"https://yakeey.com/fr-ma/achat/biens/{item['userRef']}"
    else:
        url = item.get("url") or ""

    # Images (nouveau format : on reconstruit l'URL via folder + fileName)
    images = []
    for img in (item.get("images") or []):
        if not isinstance(img, dict):
            continue
        if img.get("url"):
            images.append(img["url"])
        elif img.get("folder") and img.get("fileName"):
            images.append(f"https://yakeey-media.s3.amazonaws.com/{img['folder']}/{img['fileName']}")

    # Surface : pour les programmes, on prend areaMin comme valeur principale
    if is_program:
        surface = item.get("areaMin") or item.get("areaMax")
        chambres = item.get("roomsMin") or item.get("roomsMax")
        salles_bain = item.get("bathroomsMin") or item.get("bathroomsMax")
        statut_val = (item.get("constructionState") or {}).get("code") or ""
        tag_val = (item.get("exclusivityType") or {}).get("code")
    else:
        surface = item.get("area") or item.get("surface")
        chambres = item.get("rooms") or item.get("bedrooms") or item.get("chambres")
        salles_bain = item.get("bathrooms") or item.get("sallesBain")
        statut_val = item.get("status") or item.get("statut") or ""
        tag_val = item.get("propertyTag")

    return {
        "id": str(item.get("id") or ""),
        "ref": str(ref),
        "type": type_bien,
        "quartier": quartier,
        "district": district,
        "ville": ville,
        "prix_dh": int(price) if price else None,
        "prix_min_dh": int(item.get("priceMin")) if item.get("priceMin") else None,
        "prix_max_dh": int(item.get("priceMax")) if item.get("priceMax") else None,
        "surface_m2": surface,
        "surface_min_m2": item.get("areaMin"),
        "surface_max_m2": item.get("areaMax"),
        "surface_habitable_m2": item.get("livingArea") or item.get("habitableSurface"),
        "surface_terrain_m2": item.get("landSurface") or item.get("terrainSurface"),
        "chambres": chambres,
        "salles_bain": salles_bain,
        "salles_eau": item.get("showerRooms") or item.get("sallesEau"),
        "etage": item.get("floor") or item.get("etage"),
        "nb_etages": item.get("floors") or item.get("nbEtages"),
        "annee_construction": item.get("buildYear") or item.get("anneeConstruction"),
        "statut": statut_val,
        "vue": item.get("view"),
        "orientation": item.get("orientation"),
        "neuf": item.get("isNew") or item.get("neuf") or is_program,
        "latitude": lat,
        "longitude": lng,
        "url": url,
        "mensualite_dh": (
            item.get("monthlyPayment")
            or item.get("iBuyingMonthlyPayment")
            or item.get("monthlyInstallment")
            or item.get("mensualite")
            or None
        ),
        "frais_notaire_dh": item.get("landConservationAndRegistrationFees"),
        "cout_total_dh": item.get("estimatedProjectAmount"),
        "tag": tag_val,
        "currency": item.get("currency"),
        "listing_type": listing_type,
        "nom_programme": item.get("name") if is_program else None,
        "images": images,
        "description": item.get("description"),
        "equipements": item.get("features") or item.get("equipements") or [],
    }


def _extract_page_payload(html: str) -> dict | None:
    """
    Récupère le bloc "initialData" depuis le HTML.
    Essaie d'abord le nouveau format App Router (RSC), puis fallback
    sur l'ancien format Pages Router.
    """
    data = extract_rsc_initial_data(html)
    if data is not None:
        return data
    # Fallback ancien format
    legacy = extract_next_data(html)
    if not legacy:
        return None
    pp = legacy.get("props", {}).get("pageProps", {})
    # On essaie de reconstruire un format similaire
    items = (
        pp.get("properties")
        or pp.get("listings")
        or pp.get("results")
        or pp.get("items")
        or []
    )
    return {
        "content": items,
        "totalElements": pp.get("totalCount") or pp.get("total") or len(items),
        "totalPages": pp.get("totalPages") or 0,
        "pageSize": pp.get("perPage") or 25,
        "pageNo": pp.get("page") or 0,
    }


def _build_view_url(page: int, params: dict | None = None) -> str:
    """Construit l'URL pour une vue (page + filtres optionnels)."""
    qp = {}
    if params:
        qp.update({k: v for k, v in params.items() if v is not None})
    if page > 1:
        qp["page"] = page
    if not qp:
        return BASE_URL
    from urllib.parse import urlencode
    return f"{BASE_URL}?{urlencode(qp)}"


def discover_view(params: dict | None = None) -> tuple[int, int]:
    """
    Pour un set de filtres donne, recupere (nb_pages, total_elements)
    depuis la page 1 de cette vue.
    """
    url = _build_view_url(1, params)
    html = fetch(url)
    if not html:
        return 0, 0
    payload = _extract_page_payload(html)
    if not payload:
        return 0, 0
    total_pages = payload.get("totalPages") or 0
    total_elements = payload.get("totalElements") or 0
    if not total_pages and total_elements:
        per_page = payload.get("pageSize") or 25
        total_pages = (total_elements + per_page - 1) // per_page
    return int(total_pages), int(total_elements)


def fetch_listing_page(page: int, params: dict | None = None) -> list[dict]:
    """Recupere une page d'une vue donnee."""
    url = _build_view_url(page, params)
    html = fetch(url)
    if not html:
        return []
    payload = _extract_page_payload(html)
    if not payload:
        return []
    items = payload.get("content") or []
    out = []
    for item in items:
        parsed = parse_listing_card(item)
        if parsed and parsed.get("ref"):
            out.append(parsed)
    time.sleep(REQUEST_DELAY)
    return out


def collect_view(label: str, params: dict | None, listings: dict[str, dict]) -> int:
    """
    Aspire toutes les pages d'une vue (combinaison de filtres) et fusionne
    dans le dict global `listings`. Retourne le nombre de NOUVEAUX biens
    decouverts par cette vue (ceux non deja dans `listings`).
    """
    n_pages, n_total = discover_view(params)
    if n_pages == 0:
        # Vue vide ou filtre non reconnu : on saute silencieusement (les
        # parametres inconnus sont juste ignores par Yakeey, et la page
        # par defaut s'affiche - on ne veut pas la re-aspirer en boucle).
        return 0
    before = len(listings)
    # Limite raisonnable : si Yakeey renvoie >150 pages pour une vue,
    # c'est suspect (probablement le filtre n'a pas ete applique). On
    # capote a 150 pour eviter de boucler 200 pages identiques.
    n_pages = min(n_pages, 150)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(fetch_listing_page, p, params): p for p in range(1, n_pages + 1)}
        for fut in tqdm(as_completed(futures), total=n_pages, desc=f"  {label}", leave=False):
            for L in fut.result():
                key = L.get("id") or f"{L.get('ref')}|{L.get('type')}|{L.get('prix_dh')}"
                listings.setdefault(key, L)
    new_count = len(listings) - before
    print(f"  [+] {label:40s} | pages={n_pages:3d} total_annonce={n_total:5d} | "
          f"+{new_count:4d} nouveaux (total uniques: {len(listings)})")
    return new_count


def _probe_view(baseline_total: int, baseline_first_id: str,
                params: dict) -> tuple[str, int, str]:
    """
    Probe une vue. Retourne (kind, total, first_id) :
      kind = "FILTER"   : totalElements a change (filtre applique)
      kind = "SORT"     : totalElements identique mais premier ID different (tri applique)
      kind = "IGNORED"  : totalElements et premier ID identiques (param ignore)
    """
    url = _build_view_url(1, params)
    html = fetch(url)
    if not html:
        return "IGNORED", 0, ""
    payload = _extract_page_payload(html)
    if not payload:
        return "IGNORED", 0, ""
    total = payload.get("totalElements") or 0
    items = payload.get("content") or []
    first_id = (items[0].get("id") if items else "") or ""
    if total and total != baseline_total:
        return "FILTER", total, first_id
    if first_id and first_id != baseline_first_id:
        return "SORT", total, first_id
    return "IGNORED", total, first_id


def collect_all_listings() -> list[dict]:
    """
    Aspire les annonces depuis MULTIPLES vues (sort + filtres) et fait
    l'union, pour contourner le bug de pagination de Yakeey (qui n'expose
    qu'environ 529 IDs uniques sur les ~1250 annonces annoncees).

    Strategie :
      1) On etablit la baseline (totalElements de la vue par defaut).
      2) On collecte la vue de base.
      3) Pour chaque param de filtre candidat : on probe (1 requete) pour
         voir si totalElements change. Si oui -> on aspire toute la vue.
         Si non -> param ignore par Yakeey, on saute.
      4) Pour les tris : on aspire toujours (le tri peut changer l'overlap
         meme si totalElements reste identique).
    """
    listings: dict[str, dict] = {}

    print("\n=== ETAPE 0 : Baseline ===")
    # On lit la page 1 pour avoir baseline_total et baseline_first_id
    base_html = fetch(BASE_URL)
    base_payload = _extract_page_payload(base_html) if base_html else None
    if not base_payload:
        print("[!] Impossible de lire la page 1. Le site bloque ou a change.")
        return []
    baseline_total = int(base_payload.get("totalElements") or 0)
    baseline_pages = int(base_payload.get("totalPages") or 0)
    base_items = base_payload.get("content") or []
    baseline_first_id = (base_items[0].get("id") if base_items else "") or ""
    print(f"  Yakeey annonce {baseline_total} biens sur {baseline_pages} pages.")
    print(f"  Premier ID baseline : {baseline_first_id}")

    print("\n=== ETAPE 1 : Vue par defaut ===")
    collect_view("default", None, listings)
    after_default = len(listings)
    print(f"  -> {after_default} biens uniques apres la vue par defaut.")

    print("\n=== ETAPE 2 : Probe des parametres de tri ===")
    # On ne fait qu'1 requete par variante pour detecter ceux qui marchent.
    working_sort_keys: list[tuple[str, str]] = []  # (key_name, value)
    sort_keys_tested: set[str] = set()
    for key_name in ("sort", "sortBy", "order", "orderBy", "sortOrder"):
        for s in SORT_VARIANTS:
            if s is None:
                continue
            kind, _, first_id = _probe_view(baseline_total, baseline_first_id, {key_name: s})
            if kind in ("SORT", "FILTER"):
                working_sort_keys.append((key_name, s))
                print(f"  [+] tri actif : {key_name}={s} (1er id : {first_id[:8]})")
                sort_keys_tested.add(key_name)
        # Si aucune valeur de ce key_name n'a fonctionne, on passe au suivant
    if not working_sort_keys:
        print("  [-] Aucun parametre de tri actif detecte chez Yakeey.")

    # Pour chaque tri actif, aspirer toute la vue
    for key_name, s in working_sort_keys:
        collect_view(f"{key_name}={s}", {key_name: s}, listings)

    print("\n=== ETAPE 3 : Filtres listingType (avec probe) ===")
    lt_key_winner = None
    for t in LISTING_TYPES:
        if t is None:
            continue
        if lt_key_winner:
            collect_view(f"{lt_key_winner}={t}", {lt_key_winner: t}, listings)
            continue
        for key_name in ("listingType", "listingTypes", "type"):
            kind, total, _ = _probe_view(baseline_total, baseline_first_id, {key_name: t})
            if kind == "FILTER":
                print(f"  [+] {key_name}={t} : filtre actif (total={total})")
                collect_view(f"{key_name}={t}", {key_name: t}, listings)
                lt_key_winner = key_name
                break
        if not lt_key_winner:
            print(f"  [-] listingType={t} non filtrable")

    print("\n=== ETAPE 4 : Filtres propertyCategory (avec probe) ===")
    cat_key_winner = None
    for c in PROPERTY_CATEGORIES:
        if c is None:
            continue
        if cat_key_winner:
            collect_view(f"{cat_key_winner}={c}", {cat_key_winner: c}, listings)
            continue
        for key_name in ("propertyCategory", "propertyCategoryCodes",
                         "category", "categories", "propertyType",
                         "propertyCategories"):
            kind, total, _ = _probe_view(baseline_total, baseline_first_id, {key_name: c})
            if kind == "FILTER":
                print(f"  [+] {key_name}={c} : filtre actif (total={total})")
                collect_view(f"{key_name}={c}", {key_name: c}, listings)
                cat_key_winner = key_name
                break
        if not cat_key_winner:
            print(f"  [-] propertyCategory non filtrable (abandon de l'etape)")
            break

    print("\n=== ETAPE 5 : Tranches de prix (avec probe) ===")
    price_keys_winner = None
    for lo, hi in PRICE_BUCKETS:
        if price_keys_winner:
            kmin, kmax = price_keys_winner
            collect_view(f"{kmin}={lo}&{kmax}={hi}", {kmin: lo, kmax: hi}, listings)
            continue
        for kmin, kmax in (("priceMin", "priceMax"),
                           ("minPrice", "maxPrice"),
                           ("priceRangeMin", "priceRangeMax"),
                           ("price_min", "price_max"),
                           ("priceFrom", "priceTo")):
            kind, total, _ = _probe_view(baseline_total, baseline_first_id,
                                          {kmin: lo, kmax: hi})
            if kind == "FILTER":
                print(f"  [+] {kmin}/{kmax} : filtre actif (total={total})")
                collect_view(f"{kmin}={lo}&{kmax}={hi}",
                             {kmin: lo, kmax: hi}, listings)
                price_keys_winner = (kmin, kmax)
                break
        if not price_keys_winner:
            print(f"  [-] aucun param prix reconnu (tranches abandonnees)")
            break

    print(f"\n[*] TOTAL FINAL : {len(listings)} biens uniques.")
    print(f"    Yakeey annonce {baseline_total} biens, dont {529} sont reellement exposes")
    print(f"    par la pagination standard (bug de fenetre glissante).")
    result = list(listings.values())
    LISTINGS_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


# ============================ Detail pages ============================

def fetch_detail(url: str) -> dict | None:
    """Récupère la page détail et extrait les infos enrichies (GPS, équipements...)."""
    if not url:
        return None
    html = fetch(url)
    if not html:
        return None
    # Nouveau format App Router : "property":{...} ou "initialData":{...}
    payload = extract_rsc_initial_data(html)
    if payload and isinstance(payload, dict):
        # Sur une page détail, "initialData" peut être l'objet annonce directement
        if "content" not in payload:
            time.sleep(REQUEST_DELAY)
            return parse_listing_card(payload)
        items = payload.get("content") or []
        if items:
            time.sleep(REQUEST_DELAY)
            return parse_listing_card(items[0])
    # Fallback ancien format
    data = extract_next_data(html)
    if not data:
        return None
    page_props = data.get("props", {}).get("pageProps", {})
    detail = (
        page_props.get("property")
        or page_props.get("listing")
        or page_props.get("data")
        or {}
    )
    time.sleep(REQUEST_DELAY)
    return parse_listing_card(detail)


def enrich_with_details(listings: list[dict]) -> list[dict]:
    """
    Tente d'enrichir chaque annonce via sa page détail.
    En cas d'échec (URL non résoluble, page sans __NEXT_DATA__), on garde
    l'annonce telle quelle au lieu de la perdre.
    """
    enriched: list[dict] = []
    has_url = [L for L in listings if L.get("url")]
    no_url = [L for L in listings if not L.get("url")]
    enriched.extend(no_url)  # On ne les perd pas

    if not has_url:
        DETAILS_JSON.write_text(json.dumps(listings, ensure_ascii=False, indent=2), encoding="utf-8")
        return listings

    success = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(fetch_detail, L["url"]): L for L in has_url}
        for fut in tqdm(as_completed(futures), total=len(futures), desc="Détails"):
            base = futures[fut]
            try:
                detail = fut.result()
            except Exception:
                detail = None
            if detail:
                success += 1
                for k, v in detail.items():
                    if v not in (None, "", [], 0) or base.get(k) in (None, "", []):
                        base[k] = v
            enriched.append(base)

    if success == 0:
        print(f"[!] Aucune page détail n'a pu être enrichie ({len(has_url)} tentatives).")
        print("[!] On conserve les données des pages liste (déjà très complètes).")
    else:
        print(f"[+] {success}/{len(has_url)} pages détail enrichies.")

    DETAILS_JSON.write_text(json.dumps(enriched, ensure_ascii=False, indent=2), encoding="utf-8")
    return enriched


# ============================ Export Excel ============================

def build_excel(listings: list[dict]) -> None:
    print("[*] Génération de l'Excel...")
    if not listings:
        print("[!] Aucune annonce à exporter. Vérifie que le scraper a bien collecté des données.")
        print("[!] Astuce : supprime le dossier 'yakeey_data' pour forcer un nouveau scrape.")
        # On génère quand même un Excel vide avec juste les en-têtes pour ne pas crasher.
    listings.sort(key=lambda x: (x.get("type", ""), x.get("ref", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "Annonces"

    headers = [
        "Référence", "Type", "Quartier", "Ville",
        "Prix (DH)", "Surface (m²)", "Surface habitable", "Surface terrain",
        "Prix au m² (DH/m²)",
        "Chambres", "SDB", "Salles d'eau", "Étage", "Nb étages",
        "Année", "Statut", "Vue", "Orientation",
        "Latitude", "Longitude", "Mensualité (DH)",
        "Description", "URL",
    ]
    ws.append(headers)

    # Style en-tête
    hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="1F4E78")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = border

    # Lignes
    for i, L in enumerate(listings, start=2):
        row = [
            L.get("ref"), L.get("type"), L.get("quartier"), L.get("ville"),
            L.get("prix_dh"), L.get("surface_m2"),
            L.get("surface_habitable_m2"), L.get("surface_terrain_m2"),
            f"=IFERROR(E{i}/F{i},\"\")",
            L.get("chambres"), L.get("salles_bain"), L.get("salles_eau"),
            L.get("etage"), L.get("nb_etages"),
            L.get("annee_construction"), L.get("statut"),
            L.get("vue"), L.get("orientation"),
            L.get("latitude"), L.get("longitude"),
            L.get("mensualite_dh"),
            (L.get("description") or "")[:500],
            L.get("url"),
        ]
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
        url = L.get("url")
        if url:
            c = ws.cell(row=i, column=len(headers))
            c.hyperlink = url
            c.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    n_rows = len(listings) + 1
    has_data = len(listings) > 0

    # Formats numériques
    for r in range(2, n_rows + 1):
        ws.cell(row=r, column=5).number_format = '#,##0 "DH"'        # Prix
        ws.cell(row=r, column=6).number_format = '#,##0 "m²"'        # Surface
        ws.cell(row=r, column=7).number_format = '#,##0 "m²"'
        ws.cell(row=r, column=8).number_format = '#,##0 "m²"'
        ws.cell(row=r, column=9).number_format = '#,##0 "DH/m²"'
        ws.cell(row=r, column=21).number_format = '#,##0 "DH"'       # Mensualite

    # Largeurs colonnes
    widths = {
        1: 12, 2: 16, 3: 26, 4: 12, 5: 16, 6: 12, 7: 14, 8: 14, 9: 18,
        10: 9, 11: 7, 12: 11, 13: 8, 14: 9, 15: 9, 16: 16, 17: 18, 18: 22,
        19: 11, 20: 11, 21: 14, 22: 60, 23: 60,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "B2"

    if has_data:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n_rows}"

        # Lignes alternees
        alt = PatternFill("solid", start_color="F2F2F2")
        for r in range(2, n_rows + 1):
            if r % 2 == 0:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = alt

        # Degrade sur prix/m2
        rule = ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
        ws.conditional_formatting.add(f"I2:I{n_rows}", rule)

    # === Feuille Synthese ===
    ws2 = wb.create_sheet("Synthese")
    ws2.append(["Indicateur", "Valeur"])
    for col in range(1, 3):
        c = ws2.cell(row=1, column=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align

    if has_data:
        stats = [
            ("Source", "Yakeey.com"),
            ("URL", BASE_URL),
            ("Annonces collectees", f"=COUNTA(Annonces!A2:A{n_rows})"),
            ("Prix moyen", f"=AVERAGE(Annonces!E2:E{n_rows})"),
            ("Prix median", f"=MEDIAN(Annonces!E2:E{n_rows})"),
            ("Prix min", f"=MIN(Annonces!E2:E{n_rows})"),
            ("Prix max", f"=MAX(Annonces!E2:E{n_rows})"),
            ("Surface moyenne", f"=AVERAGE(Annonces!F2:F{n_rows})"),
            ("Prix moyen au m2", f"=AVERAGE(Annonces!I2:I{n_rows})"),
            ("Nb quartiers couverts",
             f"=SUMPRODUCT(1/COUNTIF(Annonces!C2:C{n_rows},Annonces!C2:C{n_rows}))"),
        ]
    else:
        stats = [
            ("Source", "Yakeey.com"),
            ("URL", BASE_URL),
            ("Annonces collectees", 0),
            ("Statut", "Aucune donnee collectee - verifier le scraper"),
        ]

    for r, (label, val) in enumerate(stats, start=2):
        ws2.cell(row=r, column=1, value=label).font = Font(bold=True)
        cc = ws2.cell(row=r, column=2, value=val)
        if isinstance(val, str) and val.startswith("=") and "Prix" in label:
            cc.number_format = '#,##0 "DH"'
        elif isinstance(val, str) and val.startswith("=") and "Surface" in label:
            cc.number_format = '#,##0 "m2"'
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 36

    # === Feuille Par Quartier ===
    ws3 = wb.create_sheet("Par Quartier")
    ws3.append(["Quartier", "Nb annonces", "Prix moyen", "Surface moyenne", "Prix moyen au m2"])
    for col in range(1, 6):
        c = ws3.cell(row=1, column=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
    quartiers = sorted({L.get("quartier") for L in listings if L.get("quartier")})
    for r, q in enumerate(quartiers, start=2):
        ws3.cell(row=r, column=1, value=q)
        ws3.cell(row=r, column=2, value=f"=COUNTIF(Annonces!C2:C{n_rows},A{r})")
        ws3.cell(row=r, column=3, value=f"=IFERROR(AVERAGEIF(Annonces!C2:C{n_rows},A{r},Annonces!E2:E{n_rows}),0)")
        ws3.cell(row=r, column=4, value=f"=IFERROR(AVERAGEIF(Annonces!C2:C{n_rows},A{r},Annonces!F2:F{n_rows}),0)")
        ws3.cell(row=r, column=5, value=f"=IFERROR(AVERAGEIF(Annonces!C2:C{n_rows},A{r},Annonces!I2:I{n_rows}),0)")
        ws3.cell(row=r, column=3).number_format = '#,##0 "DH"'
        ws3.cell(row=r, column=4).number_format = '#,##0 "m2"'
        ws3.cell(row=r, column=5).number_format = '#,##0 "DH/m2"'
    if quartiers:
        ws3.auto_filter.ref = f"A1:E{len(quartiers)+1}"
    for col, w in {1: 28, 2: 14, 3: 18, 4: 18, 5: 22}.items():
        ws3.column_dimensions[get_column_letter(col)].width = w

    wb.save(EXCEL_FILE)
    print(f"[+] Excel genere : {EXCEL_FILE}")


# ============================ Main ============================

def main():
    import sys
    refresh = "--refresh" in sys.argv or "-r" in sys.argv

    # Etape 1 : recuperer toutes les annonces (les pages liste contiennent
    # deja toutes les donnees structurees dont on a besoin).
    if LISTINGS_JSON.exists() and not refresh:
        print(f"[=] Chargement du cache : {LISTINGS_JSON}")
        print(f"    (Utiliser --refresh pour forcer une nouvelle collecte.)")
        listings = json.loads(LISTINGS_JSON.read_text(encoding="utf-8"))
    else:
        if refresh and LISTINGS_JSON.exists():
            print(f"[!] --refresh : suppression du cache existant.")
            LISTINGS_JSON.unlink()
        listings = collect_all_listings()

    # Etape 2 desactivee : Yakeey n'expose pas de page detail predictible.

    # Etape 3 : export Excel
    build_excel(listings)
    print(f"[OK] Termine. {len(listings)} biens dans l'Excel.")


if __name__ == "__main__":
    main()
